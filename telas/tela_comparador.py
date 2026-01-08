import os
import pandas as pd
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from utils.constantes import CORES
from utils.auxiliares import resource_path

# =====================================================
# COMPARADOR DE NOTAS
# =====================================================

class ComparadorNotasEmbed:
    def __init__(self, parent_frame, sistema_fiscal):
        self.parent_frame = parent_frame
        self.sistema_fiscal = sistema_fiscal
        self.sefaz = None
        self.sistema = None
        self.arquivo_pdf = None
        self.criar_interface()
   

    def criar_interface(self):
        # Frame principal
        main_frame = ttk.Frame(self.parent_frame, style='Main.TFrame')
        main_frame.pack(fill="both", expand=True, padx=50, pady=30)

        # Header
        header_frame = ttk.Frame(main_frame, style='Main.TFrame')
        header_frame.pack(fill="x", pady=(0, 30))

        header_container = ttk.Frame(header_frame, style='Main.TFrame')
        header_container.pack(fill="x")

        # Ícone e título
        left_header = ttk.Frame(header_container, style='Main.TFrame')
        left_header.pack(side="left")

        # ÍCONE DO HEADER
        caminho_icone = resource_path("Icones/comparador_azul.png")
        img = Image.open(caminho_icone)
        img = img.resize((32, 32), Image.LANCZOS)
        self.icon_header = ImageTk.PhotoImage(img)  # manter referência!

        ttk.Label(
            left_header,
            image=self.icon_header,
            background=CORES['bg_main']
        ).pack(side="left", padx=(0, 15))

        title_frame = ttk.Frame(left_header, style='Main.TFrame')
        title_frame.pack(side="left")

        ttk.Label(
            title_frame,
            text="Comparador de Notas Fiscais",
            font=('Segoe UI', 18, 'bold'),
            background=CORES['bg_main'],
            foreground=CORES['text_dark']
        ).pack(anchor="w")
        
        ttk.Label(
            title_frame,
            text="Compare notas fiscais entre SEFAZ e Sistema interno",
            font=('Segoe UI', 9),
            background=CORES['bg_main'],
            foreground=CORES['text_light']
        ).pack(anchor="w")
        
        # Card de conteúdo
        content_container = ttk.Frame(main_frame, style='Main.TFrame')
        content_container.pack(fill="both", expand=True)

        center_frame = ttk.Frame(content_container, style='Main.TFrame')
        center_frame.pack(expand=True)

        card = ttk.Frame(center_frame, style='Card.TFrame', padding=40)
        card.pack(fill="both", expand=True, ipadx=100)

        ttk.Label(
            card,
            text="Configuração da Comparação",
            font=('Segoe UI', 14, 'bold'),
            background=CORES['bg_card'],
            foreground=CORES['primary']
        ).pack(anchor="w", pady=(0, 25))
        
        # Arquivo SEFAZ
        self.criar_card_arquivo(
            card,
            "1. Arquivo SEFAZ",
            "Planilha com dados da Secretaria da Fazenda",
            "sefaz",
            self.sel_sefaz
        )
        
        # Arquivo Sistema
        self.criar_card_arquivo(
            card,
            "2. Arquivo Sistema",
            "Planilha com dados do sistema interno",
            "sistema",
            self.sel_sistema
        )
        
        # Status
        self.status_frame = ttk.Frame(card, style='Card.TFrame')
        self.status_frame.pack(fill="x", pady=(15, 0))
        
        self.status_label = ttk.Label(
            self.status_frame,
            text="",
            style='Subtitle.TLabel',
            wraplength=550
        )
        self.status_label.pack(anchor="w")
        
        # Botão comparar
        btn_frame = ttk.Frame(card, style='Card.TFrame')
        btn_frame.pack(fill="x", pady=(20, 0))
        
        ttk.Button(
            btn_frame,
            text="▶ Comparar e Exportar Resultado",
            style='Primary.TButton',
            command=self.comparar
        ).pack(fill="x")

    def criar_card_arquivo(self, parent, titulo, descricao, tipo, comando):
        arquivo_frame = ttk.Frame(parent, style='Card.TFrame')
        arquivo_frame.pack(fill="x", pady=(0, 15))
        
        # Cabeçalho do card
        header = ttk.Frame(arquivo_frame, style='Card.TFrame')
        header.pack(fill="x")
        
        ttk.Label(
            header,
            text=titulo,
            style='Title.TLabel'
        ).pack(side="left")
        
        # Badge de status
        status_badge = ttk.Label(
            header,
            text="Não selecionado",
            font=('Segoe UI', 8),
            background='#fee2e2',
            foreground='#991b1b',
            padding=(8, 2)
        )
        status_badge.pack(side="right")
        setattr(self, f'badge_{tipo}', status_badge)
        
        ttk.Label(
            arquivo_frame,
            text=descricao,
            style='Subtitle.TLabel'
        ).pack(anchor="w", pady=(2, 8))
        
        # Caminho do arquivo
        path_label = ttk.Label(
            arquivo_frame,
            text="Nenhum arquivo selecionado",
            font=('Segoe UI', 8, 'italic'),
            background=CORES['bg_card'],
            foreground=CORES['text_light']
        )
        path_label.pack(anchor="w", pady=(0, 8))
        setattr(self, f'path_{tipo}', path_label)
        
        ttk.Button(
            arquivo_frame,
            text="📁 Selecionar Arquivo Excel",
            style='Secondary.TButton',
            command=comando
        ).pack(anchor="w")

    def atualizar_status(self, tipo, caminho):
        badge = getattr(self, f'badge_{tipo}')
        path_label = getattr(self, f'path_{tipo}')
        
        badge.config(
            text="✓ Selecionado",
            background='#dcfce7',
            foreground='#166534'
        )
        
        # Mostrar apenas o nome do arquivo
        nome_arquivo = os.path.basename(caminho)
        path_label.config(text=f"📄 {nome_arquivo}")
        
        # Atualizar status geral
        if self.sefaz and self.sistema:
            self.status_label.config(
                text="✓ Todos os arquivos selecionados. Pronto para comparar!",
                foreground=CORES['success']
            )

    def sel_sefaz(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo SEFAZ",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.sefaz = arquivo
            self.atualizar_status('sefaz', arquivo)

    def sel_sistema(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo Sistema",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.sistema = arquivo
            self.atualizar_status('sistema', arquivo)

    def comparar(self):
        if not self.sefaz or not self.sistema:
            messagebox.showwarning(
                "Arquivos Obrigatórios",
                "Por favor, selecione ambos os arquivos (SEFAZ e Sistema) antes de continuar."
            )
            return

        try:
            self.status_label.config(
                text="⏳ Processando arquivos...",
                foreground=CORES['text_light']
            )
            self.parent_frame.update_idletasks()

            # Ler arquivos
            sefaz = pd.read_excel(self.sefaz)
            sistema = pd.read_excel(self.sistema)

            # ===== VERIFICAR COLUNAS =====
            print("\n=== DEBUG: COLUNAS ENCONTRADAS ===")
            print(f"SEFAZ colunas: {sefaz.columns.tolist()}")
            print(f"Sistema colunas: {sistema.columns.tolist()}")

            # ===== NORMALIZAÇÃO =====
            sefaz_aux = sefaz.copy()
            sistema_aux = sistema.copy()

            # ===== NORMALIZAR NÚMERO =====
            # Verificar qual coluna existe no SEFAZ
            col_numero_sefaz = None
            for col in ['Número', 'Numero', 'número', 'numero']:
                if col in sefaz_aux.columns:
                    col_numero_sefaz = col
                    break
            
            if not col_numero_sefaz:
                raise ValueError(f"Coluna de número não encontrada no SEFAZ. Colunas disponíveis: {sefaz_aux.columns.tolist()}")

            sefaz_aux["numero"] = (
                sefaz_aux[col_numero_sefaz]
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .str.lstrip("0")
            )

            # Verificar qual coluna existe no Sistema
            col_numero_sistema = None
            for col in ['Numero', 'Número', 'numero', 'número']:
                if col in sistema_aux.columns:
                    col_numero_sistema = col
                    break
            
            if not col_numero_sistema:
                raise ValueError(f"Coluna de número não encontrada no Sistema. Colunas disponíveis: {sistema_aux.columns.tolist()}")

            sistema_aux["numero"] = (
                sistema_aux[col_numero_sistema]
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .str.lstrip("0")
            )

            # ===== NORMALIZAR VALOR =====
            def limpar_valor_br(valor):
                """Converte valor brasileiro (R$ 19.454,00) para float"""
                if pd.isna(valor):
                    return 0.0
                valor_str = str(valor).strip()
                # Remove R$ e espaços
                valor_str = valor_str.replace("R$", "").strip()
                # Remove espaços em branco
                valor_str = valor_str.replace(" ", "")
                # Substitui vírgula por ponto ANTES de remover pontos dos milhares
                valor_str = valor_str.replace(",", ".")
                # Agora remove TODOS os pontos EXCETO o último (que é o decimal)
                partes = valor_str.split(".")
                if len(partes) > 1:
                    # Junta tudo antes do último ponto + último ponto + último valor
                    valor_str = "".join(partes[:-1]) + "." + partes[-1]
                try:
                    return float(valor_str)
                except:
                    return 0.0

            # Verificar coluna de valor SEFAZ
            col_valor_sefaz = None
            for col in ['Valor Total', 'Valor', 'valor', 'valor total']:
                if col in sefaz_aux.columns:
                    col_valor_sefaz = col
                    break
            
            if not col_valor_sefaz:
                raise ValueError(f"Coluna de valor não encontrada no SEFAZ. Colunas disponíveis: {sefaz_aux.columns.tolist()}")

            sefaz_aux["valor"] = sefaz_aux[col_valor_sefaz].apply(limpar_valor_br).round(2)

            # Verificar coluna de valor Sistema
            col_valor_sistema = None
            for col in ['Valor', 'valor', 'Valor Total', 'valor total']:
                if col in sistema_aux.columns:
                    col_valor_sistema = col
                    break
            
            if not col_valor_sistema:
                raise ValueError(f"Coluna de valor não encontrada no Sistema. Colunas disponíveis: {sistema_aux.columns.tolist()}")

            sistema_aux["valor"] = pd.to_numeric(
                sistema_aux[col_valor_sistema], errors="coerce"
            ).fillna(0.0).round(2)

            # Valor padronizado (string)
            sefaz_aux["valor_str"] = sefaz_aux["valor"].apply(lambda x: f"{x:.2f}")
            sistema_aux["valor_str"] = sistema_aux["valor"].apply(lambda x: f"{x:.2f}")

            # ===== DEBUG: PRIMEIRAS LINHAS =====
            print("\n=== DEBUG: PRIMEIRAS 3 LINHAS PROCESSADAS ===")
            print("\nSEFAZ:")
            print(sefaz_aux[["numero", "valor", "valor_str"]].head(3))
            print("\nSISTEMA:")
            print(sistema_aux[["numero", "valor", "valor_str"]].head(3))

            # ===== CNPJ OPCIONAL =====
            def normalizar_cnpj(df, coluna):
                if coluna in df.columns:
                    return (
                        df[coluna]
                        .astype(str)
                        .str.replace(r"\D", "", regex=True)
                        .str.zfill(14)
                    )
                return None

            sefaz_cnpj = normalizar_cnpj(sefaz_aux, "CNPJ")
            sistema_cnpj = normalizar_cnpj(sistema_aux, "CNPJ")

            # ===== CHAVES =====
            if sefaz_cnpj is not None and sistema_cnpj is not None:
                sefaz_aux["chave"] = (
                    sefaz_aux["numero"] + "|" +
                    sefaz_cnpj + "|" +
                    sefaz_aux["valor_str"]
                )

                sistema_chaves = set(
                    sistema_aux["numero"] + "|" +
                    sistema_cnpj + "|" +
                    sistema_aux["valor_str"]
                )
            else:
                sefaz_aux["chave"] = (
                    sefaz_aux["numero"] + "|" +
                    sefaz_aux["valor_str"]
                )

                sistema_chaves = set(
                    sistema_aux["numero"] + "|" +
                    sistema_aux["valor_str"]
                )

            # ===== DEBUG: CHAVES =====
            print("\n=== DEBUG: CHAVES GERADAS ===")
            print(f"\nPrimeiras 3 chaves SEFAZ:")
            for chave in list(sefaz_aux["chave"].head(3)):
                print(f"  '{chave}'")
            
            print(f"\nTodas as chaves SISTEMA ({len(sistema_chaves)} total):")
            for chave in list(sistema_chaves):
                print(f"  '{chave}'")

            # ===== STATUS =====
            sefaz["Status"] = sefaz_aux["chave"].apply(
                lambda x: "LANÇADO" if x in sistema_chaves else "NÃO LANÇADO"
            )

            # ===== DEBUG: RESULTADO =====
            print("\n=== DEBUG: RESULTADO DA COMPARAÇÃO ===")
            print(sefaz[["Número" if "Número" in sefaz.columns else col_numero_sefaz, "Status"]].head(10))

            # ===== EXPORTAÇÃO =====
            saida = filedialog.asksaveasfilename(
                title="Salvar resultado da comparação",
                defaultextension=".xlsx",
                filetypes=[("Arquivo Excel", "*.xlsx")]
            )

            if not saida:
                self.status_label.config(text="")
                return

            self.status_label.config(text="💾 Salvando resultado...")
            self.parent_frame.update_idletasks()

            sefaz.to_excel(saida, index=False)

            # ===== FORMATAÇÃO =====
            wb = load_workbook(saida)
            ws = wb.active

            col_status = list(ws[1]).index(
                next(c for c in ws[1] if c.value == "Status")
            ) + 1

            verde = PatternFill("solid", fgColor="C6EFCE")
            vermelho = PatternFill("solid", fgColor="F4CCCC")

            for r in range(2, ws.max_row + 1):
                cor = verde if ws.cell(r, col_status).value == "LANÇADO" else vermelho
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = cor

            wb.save(saida)

            # ===== ESTATÍSTICAS =====
            total = len(sefaz)
            lancados = len(sefaz[sefaz["Status"] == "LANÇADO"])
            nao_lancados = total - lancados

            messagebox.showinfo(
                "✓ Comparação Concluída",
                f"Resultado exportado com sucesso!\n\n"
                f"Total de notas: {total}\n"
                f"Lançadas: {lancados}\n"
                f"Não lançadas: {nao_lancados}"
            )

            self.status_label.config(text="")

        except Exception as e:
            self.status_label.config(text="")
            import traceback
            erro_completo = traceback.format_exc()
            print(f"\n=== ERRO COMPLETO ===\n{erro_completo}")
            messagebox.showerror(
                "Erro no Processamento",
                f"Ocorreu um erro ao comparar os arquivos:\n\n{str(e)}"
            )