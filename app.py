import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import os
import pandas as pd
from PyPDF2 import PdfMerger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk
import re
import sys
import pdfplumber
import sqlite3
import hashlib

VERSAO_ATUAL = "1.0"
CAMINHO_EXE_ATUALIZADO = r"T:\MTSistem\app\MTSistem.exe"

DB_DIR = r"T:\MTSistem\db"
DB_PATH = os.path.join(DB_DIR, "sistemafiscal.db")


MODULOS = {
    "abrir_extrator": "Extrator TXT → Excel",
    "abrir_comparador": "Comparador SEFAZ x Sistema",
    "abrir_triagem": "Triagem SPED → PDFs",
    "abrir_extrator_pdf": "Extrator PDF → Excel"
}


def garantir_banco():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    return conn


class UsuarioDAO:
    def __init__(self):
        self.conn = garantir_banco()
        self.criar_tabelas()
        self.criar_admin_inicial()
        self.criar_configuracoes_padrao()

    # ==========================
    # TABELAS
    # ==========================
    def criar_tabelas(self):
        cur = self.conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE,
                senha TEXT NOT NULL,
                admin INTEGER DEFAULT 0
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS permissoes (
                usuario_id INTEGER,
                modulo TEXT,
                permitido INTEGER DEFAULT 0,
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS configuracoes (
                chave TEXT PRIMARY KEY,
                valor TEXT NOT NULL
            )
        """)


        self.conn.commit()

    # ==========================
    # Configurações padrão
    # ==========================
    def criar_configuracoes_padrao(self):
        configs = {
            "versao_atual": VERSAO_ATUAL,
            "atualizacao_liberada": "NAO",
            "sistema_bloqueado": "NAO",
            "mensagem_update": "Atualização disponível",
            "exe_atualizacao": CAMINHO_EXE_ATUALIZADO
        }

        cur = self.conn.cursor()

        for chave, valor in configs.items():
            cur.execute("""
                INSERT INTO configuracoes (chave, valor)
                VALUES (?, ?)
                ON CONFLICT(chave) DO NOTHING
            """, (chave, valor))

        self.conn.commit()


    def get_config(self, chave, default=None):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT valor FROM configuracoes WHERE chave=?",
            (chave,)
        )
        row = cur.fetchone()
        return row[0] if row else default

    
    def set_config(self, chave, valor):
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO configuracoes (chave, valor)
            VALUES (?, ?)
            ON CONFLICT(chave)
            DO UPDATE SET valor = excluded.valor
        """, (chave, str(valor)))
        self.conn.commit()

    def get_versao_atual_sistema(self):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT valor FROM configuracoes WHERE chave = 'versao_atual'"
        )
        row = cur.fetchone()
        return row[0] if row else None

    # ==========================
    # HASH
    # ==========================
    def hash_senha(self, senha):
        return hashlib.sha256(senha.encode("utf-8")).hexdigest()

    # ==========================
    # ADMIN PADRÃO
    # ==========================
    def criar_admin_inicial(self):
        cur = self.conn.cursor()

        # Verifica se já existe algum admin
        cur.execute("SELECT COUNT(*) FROM usuarios WHERE admin = 1")
        existe_admin = cur.fetchone()[0]

        if existe_admin == 0:
            senha_hash = self.hash_senha("123456")

            cur.execute("""
                INSERT INTO usuarios (nome, senha, admin)
                VALUES (?, ?, 1)
            """, ("admin", senha_hash))

            admin_id = cur.lastrowid

            # Concede todas as permissões
            modulos = [
                "abrir_extrator",
                "abrir_comparador",
                "abrir_triagem",
                "abrir_extrator_pdf",
                "usuarios_admin"
            ]

            for modulo in modulos:
                cur.execute("""
                    INSERT INTO permissoes (usuario_id, modulo, permitido)
                    VALUES (?, ?, 1)
                """, (admin_id, modulo))

            self.conn.commit()

    # ==========================
    # AUTENTICAÇÃO
    # ==========================
    def autenticar(self, nome, senha):
        cur = self.conn.cursor()

        # 1️⃣ Valida usuário e senha
        cur.execute(
            "SELECT id, admin FROM usuarios WHERE nome=? AND senha=?",
            (nome, self.hash_senha(senha))
        )
        usuario = cur.fetchone()

        if not usuario:
            return None  # usuário ou senha inválidos

        usuario_id, is_admin = usuario

        # 2️⃣ Verifica se o sistema está bloqueado
        cur.execute(
            "SELECT valor FROM configuracoes WHERE chave='sistema_bloqueado'"
        )
        row = cur.fetchone()

        sistema_bloqueado = row and row[0] == "SIM"

        # 3️⃣ Se bloqueado e NÃO for admin → bloqueia acesso
        if sistema_bloqueado and not is_admin:
            return "BLOQUEADO"

        # 4️⃣ Login permitido
        return usuario_id, is_admin



    def listar_usuarios(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, nome, admin FROM usuarios")
        return cur.fetchall()

    def permissoes_usuario(self, usuario_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT modulo FROM permissoes WHERE usuario_id=? AND permitido=1",
            (usuario_id,)
        )
        return [r[0] for r in cur.fetchall()]
    
    def is_admin(self, usuario_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT admin FROM usuarios WHERE id=?",
            (usuario_id,)
        )
        row = cur.fetchone()
        return row and row[0] == 1
    
    def criar_usuario(self, nome, senha, admin):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO usuarios (nome, senha, admin) VALUES (?, ?, ?)",
            (nome, self.hash_senha(senha), admin)
        )
        self.conn.commit()
        return cur.lastrowid



    def salvar_permissoes(self, usuario_id, permissoes):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM permissoes WHERE usuario_id=?", (usuario_id,))
        
        for modulo, permitido in permissoes.items():
            cur.execute(
                "INSERT INTO permissoes (usuario_id, modulo, permitido) VALUES (?, ?, ?)",
                (usuario_id, modulo, int(permitido))
            )
        self.conn.commit()

    def buscar_usuario(self, usuario_id):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id, nome, admin FROM usuarios WHERE id=?",
            (usuario_id,)
        )
        return cur.fetchone()


    def atualizar_usuario(self, usuario_id, nome, admin):
        cur = self.conn.cursor()
        cur.execute(
            "UPDATE usuarios SET nome=?, admin=? WHERE id=?",
            (nome, admin, usuario_id)
        )
        self.conn.commit()



    def atualizar_senha(self, usuario_id, senha):
        cur = self.conn.cursor()
        cur.execute(
            "UPDATE usuarios SET senha=? WHERE id=?",
            (self.hash_senha(senha), usuario_id)
        )
        self.conn.commit()


    def excluir_usuario(self, usuario_id):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM permissoes WHERE usuario_id=?", (usuario_id,))
        cur.execute("DELETE FROM usuarios WHERE id=?", (usuario_id,))
        self.conn.commit()
    
    def usuario_admin(self, usuario_id):
        cur = self.conn.cursor()
        cur.execute("SELECT admin FROM usuarios WHERE id=?", (usuario_id,))
        row = cur.fetchone()
        return bool(row and row[0])

    


# =====================================================
# Tela de Login
# =====================================================

class TelaLogin:
    def __init__(self, root, versao_remota):
        self.root = root
        self.versao_remota = versao_remota
        self.dao = UsuarioDAO()


        self.root.title("MTSistem - Login")
        self.root.configure(bg=CORES['bg_main'])

        # Tamanho responsivo
        largura = int(self.root.winfo_screenwidth() * 0.35)
        altura = int(self.root.winfo_screenheight() * 0.55)
        self.root.geometry(f"{largura}x{altura}")
        self.root.minsize(500, 450)
        self.root.resizable(True, True)

        # ÍCONE
        caminho_icone = resource_path("Icones/logo.ico")
        self.root.iconbitmap(caminho_icone)

        self.centralizar_janela()
        configurar_estilo()
        self.criar_interface()

    def centralizar_janela(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def criar_interface(self):
        # =========================
        # FRAME PRINCIPAL
        # =========================
        main_frame = ttk.Frame(self.root, style="Main.TFrame", padding=30)
        main_frame.pack(fill="both", expand=True)

        # =========================
        # LOGO
        # =========================
        caminho_logo = resource_path("Icones/logo.png")
        img = Image.open(caminho_logo)
        img = img.resize((80, 80), Image.LANCZOS)
        self.logo_img = ImageTk.PhotoImage(img)

        ttk.Label(
            main_frame,
            image=self.logo_img,
            background=CORES['bg_main']
        ).pack(pady=(0, 15))

        ttk.Label(
            main_frame,
            text="Sistema Fiscal",
            font=('Segoe UI', 20, 'bold'),
            foreground=CORES['text_dark'],
            background=CORES['bg_main']
        ).pack()

        ttk.Label(
            main_frame,
            text="Selecione o usuário para entrar",
            font=('Segoe UI', 10),
            foreground=CORES['text_light'],
            background=CORES['bg_main']
        ).pack(pady=(5, 25))

        # =========================
        # CARD DE USUÁRIOS
        # =========================
        card = ttk.Frame(main_frame, style="Card.TFrame", padding=25)
        card.pack(fill="x")

        for user_id, nome, admin in self.dao.listar_usuarios():
            texto = f"{nome} {'(Admin)' if admin else ''}"

            btn = ttk.Button(
                card,
                text=texto,
                command=lambda i=user_id, n=nome: self.pedir_senha(i, n)
            )
            btn.pack(fill="x", pady=6)

        # =========================
        # RODAPÉ
        # =========================
        ttk.Label(
            main_frame,
            text="© MTSistem • Desenvolvido por Miquéias Teles",
            font=('Segoe UI', 8),
            foreground=CORES['text_light'],
            background=CORES['bg_main']
        ).pack(pady=(30, 0))

    def pedir_senha(self, usuario_id, nome):
        senha = simpledialog.askstring(
            "Senha",
            f"Digite a senha do usuário {nome}:",
            show="*",
            parent=self.root
        )

        if not senha:
            return

        auth = self.dao.autenticar(nome, senha)

        # ❌ Usuário ou senha inválidos
        if auth is None:
            messagebox.showerror(
                "Erro",
                "Usuário ou senha inválidos.",
                parent=self.root
            )
            return

        # 🔒 Sistema bloqueado
        if auth == "BLOQUEADO":
            messagebox.showwarning(
                "Sistema bloqueado",
                "O acesso está temporariamente bloqueado.\n\n"
                "Consulte o administrador.",
                parent=self.root
            )
            return

        # ✅ Login permitido
        usuario_id, is_admin = auth

        self.root.destroy()
        root = tk.Tk()
        SistemaFiscal(
            root,
            usuario_id=usuario_id,
            usuario_nome=nome,
            versao_remota=self.versao_remota
        )
        root.mainloop()



# =====================================================
# Tela de Cadastro de Usuários e Permissões
# =====================================================
class TelaUsuariosAdmin:
    def __init__(self, parent):
        self.dao = UsuarioDAO()
        self.usuario_atual_id = None
        self.vars_permissoes = {}

        self.janela = tk.Toplevel(parent)
        self.janela.title("Usuários e Permissões")
        self.janela.geometry("900x600")
        self.janela.configure(bg=CORES['bg_main'])

        # ÍCONE
        caminho_icone = resource_path("Icones/logo.ico")
        self.janela.iconbitmap(caminho_icone)

        self.centralizar_janela()
        self.criar_interface()
        self.carregar_usuarios()

    def centralizar_janela(self):
        self.janela.update_idletasks()
        width = self.janela.winfo_width()
        height = self.janela.winfo_height()
        x = (self.janela.winfo_screenwidth() // 2) - (width // 2)
        y = (self.janela.winfo_screenheight() // 2) - (height // 2)
        self.janela.geometry(f"{width}x{height}+{x}+{y}")    

    def criar_interface(self):
        container = ttk.Frame(self.janela, padding=20)
        container.pack(fill="both", expand=True)

        # ======================
        # LADO ESQUERDO – LISTA
        # ======================
        left = ttk.Frame(container)
        left.pack(side="left", fill="y", padx=(0, 20))

        ttk.Label(
            left,
            text="Usuários",
            font=('Segoe UI', 14, 'bold')
        ).pack(anchor="w")

        self.lista = tk.Listbox(left, width=30, height=25)
        self.lista.pack(pady=10)
        self.lista.bind("<<ListboxSelect>>", self.selecionar_usuario)

        ttk.Button(
            left,
            text="Novo Usuário",
            command=self.novo_usuario
        ).pack(fill="x", pady=(5, 0))

        ttk.Button(
            left,
            text="Excluir Usuário",
            command=self.excluir_usuario
        ).pack(fill="x", pady=5)

        # ======================
        # LADO DIREITO – FORM
        # ======================
        right = ttk.Frame(container)
        right.pack(side="left", fill="both", expand=True)

        ttk.Label(
            right,
            text="Dados do Usuário",
            font=('Segoe UI', 14, 'bold')
        ).pack(anchor="w")

        form = ttk.Frame(right)
        form.pack(fill="x", pady=10)

        ttk.Label(form, text="Nome").grid(row=0, column=0, sticky="w")
        self.entry_nome = ttk.Entry(form)
        self.entry_nome.grid(row=0, column=1, sticky="ew", padx=10)

        ttk.Label(form, text="Nova senha").grid(row=1, column=0, sticky="w")
        self.entry_senha = ttk.Entry(form, show="*")
        self.entry_senha.grid(row=1, column=1, sticky="ew", padx=10)

        self.var_admin = tk.IntVar()
        ttk.Checkbutton(
            form,
            text="Administrador",
            variable=self.var_admin
        ).grid(row=2, column=1, sticky="w", pady=5)

        form.columnconfigure(1, weight=1)

        # ======================
        # PERMISSÕES
        # ======================
        ttk.Label(
            right,
            text="Permissões",
            font=('Segoe UI', 12, 'bold')
        ).pack(anchor="w", pady=(15, 5))

        perms = ttk.Frame(right)
        perms.pack(anchor="w")

        for modulo, label in MODULOS.items():
            var = tk.IntVar()
            self.vars_permissoes[modulo] = var
            ttk.Checkbutton(
                perms,
                text=label,
                variable=var
            ).pack(anchor="w")

        ttk.Button(
            right,
            text="Salvar alterações",
            command=self.salvar
        ).pack(pady=20)

    # ======================
    # MÉTODOS
    # ======================
    def carregar_usuarios(self):
        self.lista.delete(0, tk.END)
        self.usuarios = self.dao.listar_usuarios()
        for uid, nome, admin in self.usuarios:
            sufixo = " (Admin)" if admin else ""
            self.lista.insert(tk.END, f"{nome}{sufixo}")

    def selecionar_usuario(self, event):
        idx = self.lista.curselection()
        if not idx:
            return

        self.usuario_atual_id = self.usuarios[idx[0]][0]
        dados = self.dao.buscar_usuario(self.usuario_atual_id)

        self.entry_nome.delete(0, tk.END)
        self.entry_nome.insert(0, dados[1])
        self.var_admin.set(dados[2])

        self.entry_senha.delete(0, tk.END)

        permissoes = self.dao.permissoes_usuario(self.usuario_atual_id)
        for modulo, var in self.vars_permissoes.items():
            var.set(1 if modulo in permissoes else 0)

    def novo_usuario(self):
        self.usuario_atual_id = None
        self.entry_nome.delete(0, tk.END)
        self.entry_senha.delete(0, tk.END)
        self.var_admin.set(0)
        for var in self.vars_permissoes.values():
            var.set(0)
    
    def salvar(self):
        nome = self.entry_nome.get().strip()
        senha = self.entry_senha.get().strip()
        admin = self.var_admin.get()

        if not nome:
            messagebox.showerror("Erro", "Nome é obrigatório")
            return

        if self.usuario_atual_id:
            self.dao.atualizar_usuario(self.usuario_atual_id, nome, admin)

            if senha:
                self.dao.atualizar_senha(self.usuario_atual_id, senha)

            usuario_id = self.usuario_atual_id
        else:
            if not senha:
                messagebox.showerror("Erro", "Senha obrigatória para novo usuário")
                return

            usuario_id = self.dao.criar_usuario(nome, senha, admin)

        permissoes = {
            modulo: var.get()
            for modulo, var in self.vars_permissoes.items()
        }
        self.dao.salvar_permissoes(usuario_id, permissoes)

        self.carregar_usuarios()
        messagebox.showinfo("Sucesso", "Usuário salvo com sucesso")

    def excluir_usuario(self):
        if not self.usuario_atual_id:
            return

        if messagebox.askyesno(
            "Confirmar",
            "Deseja realmente excluir este usuário?"
        ):
            self.dao.excluir_usuario(self.usuario_atual_id)
            self.novo_usuario()
            self.carregar_usuarios()


# =====================================================
# Funções auxiliares
# =====================================================

def resource_path(rel_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, rel_path)
    return os.path.join(os.path.abspath("."), rel_path)



def pasta_dados_app(nome_app="MTSistem"):
    base = os.path.join(os.path.expanduser("~"), "Documents", nome_app)
    os.makedirs(base, exist_ok=True)
    return base

def verificar_versao_no_startup():
    dao = UsuarioDAO()

    versao_remota = dao.get_config("versao_atual")

    if not versao_remota:
        messagebox.showerror(
            "Erro crítico",
            "Versão do sistema não configurada no banco de dados.\n\n"
            "O sistema não pode ser iniciado."
        )
        sys.exit()

    return versao_remota


def atualizacao_liberada(dao: UsuarioDAO):
    try:
        return dao.get_config("atualizacao_liberada") == "SIM"
    except Exception:
        return False
    

def sistema_esta_desatualizado(dao):
    versao_banco = dao.get_config("versao_atual")
    return str(versao_banco).strip() != str(VERSAO_ATUAL).strip()



def executar_atualizacao(dao: UsuarioDAO):
    caminho_exe = dao.get_config("exe_atualizacao")

    if not caminho_exe or not os.path.exists(caminho_exe):
        messagebox.showerror(
            "Atualização indisponível",
            "Arquivo de atualização não encontrado.\n\n"
            "Entre em contato com o administrador."
        )
        return

    try:
        messagebox.showinfo(
            "Atualização",
            "O sistema será fechado para aplicar a atualização."
        )

        os.startfile(caminho_exe)
        sys.exit()

    except OSError:
        messagebox.showwarning(
            "Atualização cancelada",
            "A atualização foi cancelada pelo usuário."
        )


# =====================================================
# CONFIGURAÇÕES DE ESTILO
# =====================================================

CORES = {
    'bg_card_hover': '#F5F7FA',
    'primary': '#2563eb',
    'primary_hover': '#1d4ed8',
    'secondary': '#64748b',
    'success': '#10b981',
    'danger': '#ef4444',
    'bg_main': '#f8fafc',
    'bg_card': '#ffffff',
    'text_dark': '#1e293b',
    'text_light': '#64748b',
    'border': '#e2e8f0'
}

def configurar_estilo():
    """Configura estilos modernos para os widgets ttk"""
    style = ttk.Style()
    style.theme_use('clam')
    
    # Botões primários
    style.configure(
        'Primary.TButton',
        background=CORES['primary'],
        foreground='white',
        borderwidth=0,
        focuscolor='none',
        font=('Segoe UI', 10),
        padding=(20, 10)
    )
    style.map('Primary.TButton',
        background=[('active', CORES['primary_hover'])]
    )
    
    # Botões secundários
    style.configure(
        'Secondary.TButton',
        background=CORES['bg_card'],
        foreground=CORES['text_dark'],
        borderwidth=1,
        relief='solid',
        font=('Segoe UI', 9),
        padding=(15, 8)
    )

    # Botões voltar
    style.configure(
        'Voltar.TButton',
        background=CORES['primary'],
        foreground='white',
        borderwidth=0,
        focuscolor='none',
        font=('Segoe UI', 8),
        padding=(10, 8)
    )
    style.map('Voltar.TButton',
        background=[('active', CORES['primary_hover'])]
    )
    
    # Labels
    style.configure(
        'Title.TLabel',
        background=CORES['bg_main'],
        foreground=CORES['text_dark'],
        font=('Segoe UI', 11, 'bold')
    )
    
    style.configure(
        'Subtitle.TLabel',
        background=CORES['bg_card'],
        foreground=CORES['text_light'],
        font=('Segoe UI', 9)
    )
    
    # Frames
    style.configure('Card.TFrame', background=CORES['bg_card'], relief='flat')
    style.configure('Main.TFrame', background=CORES['bg_main'])

# =====================================================
# Extração Informações TXT → Excel
# =====================================================

class ExtratorFiscalAppEmbed:
    def __init__(self, parent_frame, sistema_fiscal):
        self.parent_frame = parent_frame
        self.sistema_fiscal = sistema_fiscal
        self.arquivo_pdf = None
        self.criar_interface()

    def criar_interface(self):
        # Frame principal
        main_frame = ttk.Frame(self.parent_frame, style='Main.TFrame')
        main_frame.pack(fill="both", expand=True, padx=50, pady=30)

        # Header
        header_frame = ttk.Frame(main_frame, style='Main.TFrame')
        header_frame.pack(fill="x", pady=(0, 30))

        header_container = ttk.Frame(header_frame, style='Main.TFrame')
        header_container.pack(fill="x")

        # Ícone e título
        left_header = ttk.Frame(header_container, style='Main.TFrame')
        left_header.pack(side="left")


        caminho_icone = resource_path("Icones/txt_azul.png")
        img = Image.open(caminho_icone)
        img = img.resize((32, 32), Image.LANCZOS)
        self.icon_header = ImageTk.PhotoImage(img)  # manter referência!


        ttk.Label(
            left_header,
            image=self.icon_header,
            background=CORES['bg_main']
        ).pack(side="left", padx=(0, 15))

        title_frame = ttk.Frame(left_header, style='Main.TFrame')
        title_frame.pack(side="left")

        ttk.Label(
            title_frame,
            text="Extração TXT → Excel",
            font=('Segoe UI', 18, 'bold'),
            background=CORES['bg_main'],
            foreground=CORES['text_dark']
        ).pack(anchor="w")

        ttk.Label(
            title_frame,
            text="Extraia informações fiscais de arquivos TXT para planilhas Excel",
            font=('Segoe UI', 9),
            background=CORES['bg_main'],
            foreground=CORES['text_light']
        ).pack(anchor="w")

        # Card de conteúdo
        content_container = ttk.Frame(main_frame, style='Main.TFrame')
        content_container.pack(fill="both", expand=True)

        center_frame = ttk.Frame(content_container, style='Main.TFrame')
        center_frame.pack(expand=True)

        card = ttk.Frame(center_frame, style='Card.TFrame', padding=40)
        card.pack(fill="both", expand=True, ipadx=100)

        ttk.Label(
            card,
            text="Configuração da Extração",
            font=('Segoe UI', 14, 'bold'),
            background=CORES['bg_card'],
            foreground=CORES['primary']
        ).pack(anchor="w", pady=(0, 25))
        
        # Campos de entrada
        self.lbl_arquivo = self.criar_campo(
            card, 
            "Arquivo TXT (.txt)", 
            "Selecione o arquivo TXT para processamento",
            self.selecionar_txt
        )
        
        
        # Botão executar
        btn_frame = ttk.Frame(card, style='Card.TFrame')
        btn_frame.pack(fill="x", pady=(20, 0))
        
        ttk.Button(
            btn_frame,
            text="▶ Executar Extração e Exportar Excel",
            style='Primary.TButton',
            command=self.processar
        ).pack(fill="x")

    def criar_campo(self, parent, titulo, subtitulo, comando):
        campo_frame = ttk.Frame(parent, style='Card.TFrame')
        campo_frame.pack(fill="x", pady=(0, 20))
        
        ttk.Label(
            campo_frame,
            text=titulo,
            style='Title.TLabel'
        ).pack(anchor="w")
        
        ttk.Label(
            campo_frame,
            text=subtitulo,
            style='Subtitle.TLabel'
        ).pack(anchor="w", pady=(2, 8))
        
        input_frame = ttk.Frame(campo_frame, style='Card.TFrame')
        input_frame.pack(fill="x")
        
        entry = ttk.Entry(input_frame, font=('Segoe UI', 9))
        entry.pack(side="left", fill="x", expand=True, ipady=5)
        
        ttk.Button(
            input_frame,
            text="📁 Selecionar",
            style='Secondary.TButton',
            command=comando
        ).pack(side="right", padx=(10, 0))
        
        return entry

    def selecionar_txt(self):
        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos TXT", "*.txt")])
        if arquivo:
            self.lbl_arquivo.delete(0, tk.END)
            self.lbl_arquivo.insert(0, arquivo)
    
    def linha_valida(self, linha):
        """
        Linha válida:
        |dd/mm/aa|...
        """
        return bool(re.match(r"^\|\d{2}/\d{2}/\d{2}\|", linha))

    

    def processar(self):
        # Validação
        if not self.lbl_arquivo.get():
            messagebox.showwarning("Atenção", "Selecione um arquivo TXT.")
            return
        

        salvar_em = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )    
        
        if not salvar_em:
            return

        dados = []
        nota_atual = None  # guarda a última nota válida

        try:
            with open(self.lbl_arquivo.get(), "r", encoding="latin-1") as f:
                for linha in f:
                    linha = linha.rstrip()

                    partes = [p.strip() for p in linha.split("|")]

                    # proteção mínima
                    if len(partes) < 9:
                        continue

                    # tenta capturar valor contábil (pode existir em linhas filhas)
                    valor_txt = partes[8]
                    valor_txt = valor_txt.replace(".", "").replace(",", ".")

                    try:
                        valor = float(valor_txt)
                    except ValueError:
                        valor = 0.0

                    # linha principal: começa com |dd/mm/aa|
                    if self.linha_valida(linha):
                        numero = partes[4].lstrip("0")
                        data_emissao = partes[5]

                        nota_atual = {
                            "Numero": numero,
                            "Data de Emissao": data_emissao,
                            "Valor": valor
                        }
                        dados.append(nota_atual)

                    # linha complementar: soma ao valor da última nota
                    else:
                        if nota_atual and valor > 0:
                            nota_atual["Valor"] += valor

            if not dados:
                messagebox.showwarning(
                    "Aviso",
                    "Nenhum documento fiscal válido foi encontrado."
                )
                return

            df = pd.DataFrame(dados)

            df["Data de Emissao"] = pd.to_datetime(
                df["Data de Emissao"],
                dayfirst=True,
                errors="coerce"
            )

            df.to_excel(salvar_em, index=False)

            messagebox.showinfo(
                "Sucesso",
                f"Arquivo exportado com sucesso!\n\n"
                f"Registros extraídos: {len(df)}"
            )

        except Exception as e:
            messagebox.showerror("Erro", str(e))    

# =====================================================
# TRIAGEM SPED → PDFs
# =====================================================

class TriagemSPEDEmbed:
    def __init__(self, parent_frame, sistema_fiscal):
        self.parent_frame = parent_frame
        self.sistema_fiscal = sistema_fiscal
        self.arquivo_pdf = None
        base_dados = pasta_dados_app()
        self.pasta_padrao = os.path.join(base_dados, "Arquivos Notas PDF")
        os.makedirs(self.pasta_padrao, exist_ok=True)
        self.criar_interface()

    def criar_interface(self):
        # Frame principal
        main_frame = ttk.Frame(self.parent_frame, style='Main.TFrame')
        main_frame.pack(fill="both", expand=True, padx=50, pady=30)

        # Header
        header_frame = ttk.Frame(main_frame, style='Main.TFrame')
        header_frame.pack(fill="x", pady=(0, 30))

        header_container = ttk.Frame(header_frame, style='Main.TFrame')
        header_container.pack(fill="x")

        # Ícone e título
        left_header = ttk.Frame(header_container, style='Main.TFrame')
        left_header.pack(side="left")

        caminho_icone = resource_path("Icones/sped_azul.png")
        img = Image.open(caminho_icone)
        img = img.resize((32, 32), Image.LANCZOS)
        self.icon_header = ImageTk.PhotoImage(img)  # manter referência!


        ttk.Label(
            left_header,
            image=self.icon_header,
            background=CORES['bg_main']
        ).pack(side="left", padx=(0, 15))

        title_frame = ttk.Frame(left_header, style='Main.TFrame')
        title_frame.pack(side="left")

        ttk.Label(
            title_frame,
            text="Triagem SPED → PDFs",
            font=('Segoe UI', 18, 'bold'),
            background=CORES['bg_main'],
            foreground=CORES['text_dark']
        ).pack(anchor="w")

        ttk.Label(
            title_frame,
            text="Extraia informações do SPED e gere PDFs mesclados das notas fiscais",
            font=('Segoe UI', 9),
            background=CORES['bg_main'],
            foreground=CORES['text_light']
        ).pack(anchor="w")

        # Card de conteúdo
        content_container = ttk.Frame(main_frame, style='Main.TFrame')
        content_container.pack(fill="both", expand=True)

        center_frame = ttk.Frame(content_container, style='Main.TFrame')
        center_frame.pack(expand=True)

        card = ttk.Frame(center_frame, style='Card.TFrame', padding=40)
        card.pack(fill="both", expand=True, ipadx=100)

        ttk.Label(
            card,
            text="Configuração da Triagem SPED",
            font=('Segoe UI', 14, 'bold'),
            background=CORES['bg_card'],
            foreground=CORES['primary']
        ).pack(anchor="w", pady=(0, 25))
        
        # Campos de entrada
        self.entry_sped = self.criar_campo(
            card, 
            "Arquivo SPED (.txt)", 
            "Selecione o arquivo SPED para processamento",
            self.selecionar_sped
        )
        
        self.entry_pdfs = self.criar_campo(
            card,
            "Pasta dos PDFs",
            "Pasta contendo os arquivos PDF das notas",
            self.selecionar_pdfs
        )

        # Agora o entry EXISTE
        self.entry_pdfs.insert(0, self.pasta_padrao)

        
        self.entry_saida = self.criar_campo(
            card, 
            "Pasta de saída", 
            "Local onde os arquivos mesclados serão salvos",
            self.selecionar_saida
        )
        
        # Barra de progresso (inicialmente oculta)
        self.progress_frame = ttk.Frame(card, style='Card.TFrame')
        self.progress_frame.pack(fill="x", pady=(20, 0))
        
        self.progress_label = ttk.Label(
            self.progress_frame,
            text="",
            style='Subtitle.TLabel'
        )
        self.progress_label.pack(anchor="w", pady=(0, 5))
        
        self.progress = ttk.Progressbar(
            self.progress_frame,
            mode='indeterminate',
            length=300
        )
        
        # Botão executar
        btn_frame = ttk.Frame(card, style='Card.TFrame')
        btn_frame.pack(fill="x", pady=(20, 0))
        
        ttk.Button(
            btn_frame,
            text="▶ Executar Triagem e Gerar PDFs",
            style='Primary.TButton',
            command=self.executar
        ).pack(fill="x")

    def criar_campo(self, parent, titulo, subtitulo, comando):
        campo_frame = ttk.Frame(parent, style='Card.TFrame')
        campo_frame.pack(fill="x", pady=(0, 20))
        
        ttk.Label(
            campo_frame,
            text=titulo,
            style='Title.TLabel'
        ).pack(anchor="w")
        
        ttk.Label(
            campo_frame,
            text=subtitulo,
            style='Subtitle.TLabel'
        ).pack(anchor="w", pady=(2, 8))
        
        input_frame = ttk.Frame(campo_frame, style='Card.TFrame')
        input_frame.pack(fill="x")
        
        entry = ttk.Entry(input_frame, font=('Segoe UI', 9))
        entry.pack(side="left", fill="x", expand=True, ipady=5)
        
        ttk.Button(
            input_frame,
            text="📁 Selecionar",
            style='Secondary.TButton',
            command=comando
        ).pack(side="right", padx=(10, 0))
        
        return entry

    def selecionar_sped(self):
        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos SPED", "*.txt")])
        if arquivo:
            self.entry_sped.delete(0, tk.END)
            self.entry_sped.insert(0, arquivo)

    def selecionar_pdfs(self):
        pasta_padrao = self.pasta_padrao

        pasta = filedialog.askdirectory(
            title="Selecione a pasta dos PDFs",
            initialdir=pasta_padrao
        )

        if pasta:
            self.entry_pdfs.delete(0, tk.END)
            self.entry_pdfs.insert(0, pasta)

    def selecionar_saida(self):
        pasta = filedialog.askdirectory()
        if pasta:
            self.entry_saida.delete(0, tk.END)
            self.entry_saida.insert(0, pasta)

    def extrair_chaves(self, caminho):
        nf, cte = [], []
        with open(caminho, 'r', encoding='latin-1') as f:
            for linha in f:
                c = linha.split('|')
                if len(c) > 10:
                    if c[1] == 'C100' and c[9].isdigit():
                        nf.append(c[9])
                    if c[1] == 'D100' and c[10].isdigit():
                        cte.append(c[10])
        return nf, cte

    def mesclar(self, chaves, pasta, saida):
        merger = PdfMerger()
        for chave in chaves:
            pdf = os.path.join(pasta, f"{chave}.pdf")
            if os.path.exists(pdf):
                merger.append(pdf)
        if merger.pages:
            merger.write(saida)
            merger.close()

    def executar(self):
        # Validação
        if not self.entry_sped.get() or not self.entry_pdfs.get() or not self.entry_saida.get():
            messagebox.showwarning(
                "Campos Obrigatórios",
                "Por favor, preencha todos os campos antes de continuar."
            )
            return
        
        # Mostrar progresso
        self.progress_label.config(text="Processando arquivos SPED...")
        self.progress.pack(fill="x")
        self.progress.start(10)
        self.janela.update()
        
        try:
            nf, cte = self.extrair_chaves(self.entry_sped.get())
            
            self.progress_label.config(text=f"Mesclando {len(nf)} NF-e encontradas...")
            self.janela.update()
            self.mesclar(nf, self.entry_pdfs.get(), os.path.join(self.entry_saida.get(), "NFe_unico.pdf"))
            
            self.progress_label.config(text=f"Mesclando {len(cte)} CT-e encontradas...")
            self.janela.update()
            self.mesclar(cte, self.entry_pdfs.get(), os.path.join(self.entry_saida.get(), "CTe_unico.pdf"))
            
            self.progress.stop()
            self.progress.pack_forget()
            self.progress_label.config(text="")
            
            messagebox.showinfo(
                "✓ Concluído",
                f"PDFs gerados com sucesso!\n\n"
                f"• NF-e processadas: {len(nf)}\n"
                f"• CT-e processadas: {len(cte)}\n\n"
                f"Arquivos salvos em:\n{self.entry_saida.get()}"
            )
        except Exception as e:
            self.progress.stop()
            self.progress.pack_forget()
            self.progress_label.config(text="")
            messagebox.showerror("Erro", f"Ocorreu um erro ao processar:\n{str(e)}")


# =====================================================
# COMPARADOR DE NOTAS
# =====================================================

class ComparadorNotasEmbed:
    def __init__(self, parent_frame, sistema_fiscal):
        self.parent_frame = parent_frame
        self.sistema_fiscal = sistema_fiscal
        self.arquivo_pdf = None
        self.criar_interface()
   

    def criar_interface(self):
        # Frame principal
        main_frame = ttk.Frame(self.parent_frame, style='Main.TFrame')
        main_frame.pack(fill="both", expand=True, padx=50, pady=30)

        # Header
        header_frame = ttk.Frame(main_frame, style='Main.TFrame')
        header_frame.pack(fill="x", pady=(0, 30))

        header_container = ttk.Frame(header_frame, style='Main.TFrame')
        header_container.pack(fill="x")

        # Ícone e título
        left_header = ttk.Frame(header_container, style='Main.TFrame')
        left_header.pack(side="left")

        # ÍCONE DO HEADER
        caminho_icone = resource_path("Icones/comparador_azul.png")
        img = Image.open(caminho_icone)
        img = img.resize((32, 32), Image.LANCZOS)
        self.icon_header = ImageTk.PhotoImage(img)  # manter referência!

        ttk.Label(
            left_header,
            image=self.icon_header,
            background=CORES['bg_main']
        ).pack(side="left", padx=(0, 15))

        title_frame = ttk.Frame(left_header, style='Main.TFrame')
        title_frame.pack(side="left")

        ttk.Label(
            title_frame,
            text="Comparador de Notas Fiscais",
            font=('Segoe UI', 18, 'bold'),
            background=CORES['bg_main'],
            foreground=CORES['text_dark']
        ).pack(anchor="w")
        
        ttk.Label(
            title_frame,
            text="Compare notas fiscais entre SEFAZ e Sistema interno",
            font=('Segoe UI', 9),
            background=CORES['bg_main'],
            foreground=CORES['text_light']
        ).pack(anchor="w")
        
        # Card de conteúdo
        content_container = ttk.Frame(main_frame, style='Main.TFrame')
        content_container.pack(fill="both", expand=True)

        center_frame = ttk.Frame(content_container, style='Main.TFrame')
        center_frame.pack(expand=True)

        card = ttk.Frame(center_frame, style='Card.TFrame', padding=40)
        card.pack(fill="both", expand=True, ipadx=100)

        ttk.Label(
            card,
            text="Configuração da Comparação",
            font=('Segoe UI', 14, 'bold'),
            background=CORES['bg_card'],
            foreground=CORES['primary']
        ).pack(anchor="w", pady=(0, 25))
        
        # Arquivo SEFAZ
        self.criar_card_arquivo(
            card,
            "1. Arquivo SEFAZ",
            "Planilha com dados da Secretaria da Fazenda",
            "sefaz",
            self.sel_sefaz
        )
        
        # Arquivo Sistema
        self.criar_card_arquivo(
            card,
            "2. Arquivo Sistema",
            "Planilha com dados do sistema interno",
            "sistema",
            self.sel_sistema
        )
        
        # Status
        self.status_frame = ttk.Frame(card, style='Card.TFrame')
        self.status_frame.pack(fill="x", pady=(15, 0))
        
        self.status_label = ttk.Label(
            self.status_frame,
            text="",
            style='Subtitle.TLabel',
            wraplength=550
        )
        self.status_label.pack(anchor="w")
        
        # Botão comparar
        btn_frame = ttk.Frame(card, style='Card.TFrame')
        btn_frame.pack(fill="x", pady=(20, 0))
        
        ttk.Button(
            btn_frame,
            text="▶ Comparar e Exportar Resultado",
            style='Primary.TButton',
            command=self.comparar
        ).pack(fill="x")

    def criar_card_arquivo(self, parent, titulo, descricao, tipo, comando):
        arquivo_frame = ttk.Frame(parent, style='Card.TFrame')
        arquivo_frame.pack(fill="x", pady=(0, 15))
        
        # Cabeçalho do card
        header = ttk.Frame(arquivo_frame, style='Card.TFrame')
        header.pack(fill="x")
        
        ttk.Label(
            header,
            text=titulo,
            style='Title.TLabel'
        ).pack(side="left")
        
        # Badge de status
        status_badge = ttk.Label(
            header,
            text="Não selecionado",
            font=('Segoe UI', 8),
            background='#fee2e2',
            foreground='#991b1b',
            padding=(8, 2)
        )
        status_badge.pack(side="right")
        setattr(self, f'badge_{tipo}', status_badge)
        
        ttk.Label(
            arquivo_frame,
            text=descricao,
            style='Subtitle.TLabel'
        ).pack(anchor="w", pady=(2, 8))
        
        # Caminho do arquivo
        path_label = ttk.Label(
            arquivo_frame,
            text="Nenhum arquivo selecionado",
            font=('Segoe UI', 8, 'italic'),
            background=CORES['bg_card'],
            foreground=CORES['text_light']
        )
        path_label.pack(anchor="w", pady=(0, 8))
        setattr(self, f'path_{tipo}', path_label)
        
        ttk.Button(
            arquivo_frame,
            text="📁 Selecionar Arquivo Excel",
            style='Secondary.TButton',
            command=comando
        ).pack(anchor="w")

    def atualizar_status(self, tipo, caminho):
        badge = getattr(self, f'badge_{tipo}')
        path_label = getattr(self, f'path_{tipo}')
        
        badge.config(
            text="✓ Selecionado",
            background='#dcfce7',
            foreground='#166534'
        )
        
        # Mostrar apenas o nome do arquivo
        nome_arquivo = os.path.basename(caminho)
        path_label.config(text=f"📄 {nome_arquivo}")
        
        # Atualizar status geral
        if self.sefaz and self.sistema:
            self.status_label.config(
                text="✓ Todos os arquivos selecionados. Pronto para comparar!",
                foreground=CORES['success']
            )

    def sel_sefaz(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo SEFAZ",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.sefaz = arquivo
            self.atualizar_status('sefaz', arquivo)

    def sel_sistema(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo Sistema",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.sistema = arquivo
            self.atualizar_status('sistema', arquivo)

    def comparar(self):
        if not self.sefaz or not self.sistema:
            messagebox.showwarning(
                "Arquivos Obrigatórios",
                "Por favor, selecione ambos os arquivos (SEFAZ e Sistema) antes de continuar."
            )
            return

        try:
            self.status_label.config(
                text="⏳ Processando arquivos...",
                foreground=CORES['text_light']
            )
            self.janela.update()
            
            sefaz = pd.read_excel(self.sefaz)
            sistema = pd.read_excel(self.sistema)

            # ===== NORMALIZAÇÃO (USO INTERNO) =====
            sefaz_aux = sefaz.copy()
            sistema_aux = sistema.copy()

            sefaz_aux["numero"] = (
                sefaz_aux["Número"]
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .str.lstrip("0")
            )

            sefaz_aux["data"] = pd.to_datetime(
                sefaz_aux["Data de Emissão"], errors="coerce"
            ).dt.date

            sefaz_aux["valor"] = pd.to_numeric(
                sefaz_aux["Valor Total"], errors="coerce"
            ).round(2)

            sistema_aux["numero"] = (
                sistema_aux["Numero"]
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .str.lstrip("0")
            )

            sistema_aux["data"] = pd.to_datetime(
                sistema_aux["Data de Emissao"], errors="coerce"
            ).dt.date

            sistema_aux["valor"] = pd.to_numeric(
                sistema_aux["Valor"], errors="coerce"
            ).round(2)

            # ===== CHAVES INTERNAS =====
            sefaz_aux["chave"] = (
                sefaz_aux["numero"] + "|" +
                sefaz_aux["data"].astype(str) + "|" +
                sefaz_aux["valor"].astype(str)
            )

            sistema_chaves = set(
                sistema_aux["numero"] + "|" +
                sistema_aux["data"].astype(str) + "|" +
                sistema_aux["valor"].astype(str)
            )

            # ===== STATUS =====
            sefaz["Status"] = sefaz_aux["chave"].apply(
                lambda x: "LANÇADO" if x in sistema_chaves else "NÃO LANÇADO"
            )

            # ===== EXPORTAÇÃO =====
            saida = filedialog.asksaveasfilename(
                title="Salvar resultado da comparação",
                defaultextension=".xlsx",
                filetypes=[("Arquivo Excel", "*.xlsx")]
            )
            
            if not saida:
                self.status_label.config(text="")
                return
            
            self.status_label.config(text="💾 Salvando resultado...")
            self.janela.update()
            
            sefaz.to_excel(saida, index=False)

            # ===== FORMATAÇÃO NO EXCEL =====
            wb = load_workbook(saida)
            ws = wb.active

            col_status = list(ws[1]).index(
                next(c for c in ws[1] if c.value == "Status")
            ) + 1

            verde = PatternFill("solid", fgColor="C6EFCE")
            vermelho = PatternFill("solid", fgColor="F4CCCC")

            for r in range(2, ws.max_row + 1):
                cor = verde if ws.cell(r, col_status).value == "LANÇADO" else vermelho
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = cor

            wb.save(saida)

            self.status_label.config(
                text="✅ Comparação concluída!",
                foreground=CORES['success']
            )

            
            # Calcular estatísticas
            total = len(sefaz)
            lancados = len(sefaz[sefaz["Status"] == "LANÇADO"])
            nao_lancados = total - lancados
            
            self.status_label.config(text="")
            
            messagebox.showinfo(
                "✓ Comparação Concluída",
                f"Resultado exportado com sucesso!\n\n"
                f"Estatísticas:\n"
                f"• Total de notas: {total}\n"
                f"• Lançadas: {lancados} ({lancados/total*100:.1f}%)\n"
                f"• Não lançadas: {nao_lancados} ({nao_lancados/total*100:.1f}%)\n\n"
                f"Arquivo salvo em:\n{os.path.basename(saida)}"
            )
            
        except Exception as e:
            self.status_label.config(text="")
            messagebox.showerror(
                "Erro no Processamento",
                f"Ocorreu um erro ao comparar os arquivos:\n\n{str(e)}"
            )

# =====================================================
# Extração Informações PDF → Excel
# =====================================================
class SistemaFiscal:
    def __init__(self, root, usuario_id, usuario_nome, versao_remota):
        self.root = root
        self.usuario_id = usuario_id
        self.usuario_nome = usuario_nome
        self.dao = UsuarioDAO()
        self.versao_remota = versao_remota
        self.desatualizado = sistema_esta_desatualizado(self.dao)
        self.atualizacao_liberada = atualizacao_liberada(self.dao)
        self.usuario_admin = self.dao.usuario_admin(self.usuario_id)
        self.mensagem_update = self.dao.get_config(
            "mensagem_update",
            "⚠️ Sistema desatualizado"
        )
        root.title("MT Sistem - Sistema Fiscal")
        
        # Janela maximizada
        root.state('zoomed')
        root.minsize(1000, 600)
        root.resizable(True, True)
        root.configure(bg=CORES['bg_main'])

        # ÍCONE DA JANELA
        caminho_icone = resource_path("Icones/logo.ico")
        self.root.iconbitmap(caminho_icone)

        configurar_estilo()
        
        # Container para o conteúdo dinâmico
        self.content_area = None
        
        self.criar_interface()
    

    def abrir_extrator(self):
        self.limpar_content_area()
        ExtratorFiscalAppEmbed(self.content_area, self)

    def abrir_triagem(self):
        self.limpar_content_area()
        TriagemSPEDEmbed(self.content_area, self)

    def abrir_comparador(self):
        self.limpar_content_area()
        ComparadorNotasEmbed(self.content_area, self)

    def abrir_extrator_pdf(self):
        self.limpar_content_area()
        ExtratorFiscalPDFAppEmbed(self.content_area, self)

    def voltar_home(self):
        self.limpar_content_area()
        self.mostrar_home()

    def limpar_content_area(self):
        """Limpa a área de conteúdo"""
        if self.content_area:
            for widget in self.content_area.winfo_children():
                widget.destroy()

    def criar_interface(self):
        # =========================
        # CONTAINER PRINCIPAL
        # =========================
        main_container = ttk.Frame(self.root, style='Main.TFrame')
        main_container.pack(fill="both", expand=True)

        # =========================
        # SIDEBAR LATERAL
        # =========================
        sidebar = tk.Frame(main_container, bg=CORES['primary'], width=280)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # Logo e título no sidebar
        logo_frame = tk.Frame(sidebar, bg=CORES['primary'])
        logo_frame.pack(fill="x", pady=(30, 20))

        caminho_logo = resource_path("Icones/logo_branca.png")
        img = Image.open(caminho_logo)
        img = img.resize((60, 60), Image.LANCZOS)
        self.logo_img = ImageTk.PhotoImage(img)

        tk.Label(
            logo_frame,
            image=self.logo_img,
            bg=CORES['primary']
        ).pack()

        tk.Label(
            logo_frame,
            text="Sistema Fiscal",
            font=('Segoe UI', 16, 'bold'),
            bg=CORES['primary'],
            fg='white'
        ).pack(pady=(10, 5))

        tk.Label(
            logo_frame,
            text="MT Sistem",
            font=('Segoe UI', 9),
            bg=CORES['primary'],
            fg='white'
        ).pack()

        # Separador
        tk.Frame(sidebar, bg='white', height=1).pack(fill="x", pady=20, padx=20)

        # Verificar permissões
        dao = UsuarioDAO()
        is_admin = dao.is_admin(self.usuario_id)
        permissoes = dao.permissoes_usuario(self.usuario_id)

        # Menu de navegação
        menu_frame = tk.Frame(sidebar, bg=CORES['primary'])
        menu_frame.pack(fill="both", expand=True, pady=10)

        # Botão Home
        self.criar_menu_item(
            menu_frame,
            "Início",
            self.voltar_home,
            icone="inicio.png"
        )

        # Módulos
        if is_admin or "abrir_extrator" in permissoes:
            self.criar_menu_item(
                menu_frame,
                "Extrator TXT → Excel",
                self.abrir_extrator,
                icone="txt.png"
            )

        if is_admin or "abrir_comparador" in permissoes:
            self.criar_menu_item(
                menu_frame,
                "Comparador SEFAZ",
                self.abrir_comparador,
                icone="comparador.png"
            )

        if is_admin or "abrir_triagem" in permissoes:
            self.criar_menu_item(
                menu_frame,
                "Triagem SPED",
                self.abrir_triagem,
                icone="sped.png"
            )

        if is_admin or "abrir_extrator_pdf" in permissoes:
            self.criar_menu_item(
                menu_frame,
                "Extrator PDF → Excel",
                self.abrir_extrator_pdf,
                icone="pdf.png"
            )

        # Separador
        tk.Frame(sidebar, bg='white', height=1).pack(fill="x", pady=10, padx=20)

        # Admin
        if is_admin:
            self.criar_menu_item(
                menu_frame,
                "👥 Usuários",
                lambda: TelaUsuariosAdmin(self.root),
                is_admin_btn=True,
            )

        if is_admin:
            self.criar_menu_item(
                menu_frame,
                "⚙️ Configurações do Sistema",
                lambda: TelaConfiguracoesSistema(self.root),
                is_admin_btn=True,
            )    



        # Rodapé do sidebar
        footer_sidebar = tk.Frame(sidebar, bg=CORES['primary'])
        footer_sidebar.pack(fill="x", side="bottom", pady=20)

        # Info do usuário
        tk.Label(
            footer_sidebar,
            text=f"👤 {self.usuario_nome}",
            font=('Segoe UI', 9),
            bg=CORES['primary'],
            fg='white'
        ).pack(pady=(0, 10))

        # Botão sair
        btn_sair = tk.Button(
            footer_sidebar,
            text="🚪 Sair",
            font=('Segoe UI', 9),
            bg='white',
            fg=CORES['primary'],
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=8,
            command=self.sair
        )
        btn_sair.pack()

        def sair_hover_enter(e):
            btn_sair.config(bg=CORES['bg_card_hover'])
        
        def sair_hover_leave(e):
            btn_sair.config(bg='white')

        btn_sair.bind("<Enter>", sair_hover_enter)
        btn_sair.bind("<Leave>", sair_hover_leave)

        # =========================
        # ÁREA DE CONTEÚDO
        # =========================
        self.content_area = ttk.Frame(main_container, style='Main.TFrame')
        self.content_area.pack(side="right", fill="both", expand=True)

        # Mostrar tela inicial
        self.mostrar_home()

    def criar_menu_item(self, parent, texto, comando, icone=None, is_admin_btn=False):
        btn_frame = tk.Frame(parent, bg=CORES['primary'])
        btn_frame.pack(fill="x", padx=15, pady=5)

        btn = tk.Frame(btn_frame, bg=CORES['primary'], cursor='hand2')
        btn.pack(fill="x", pady=2)

        inner = tk.Frame(btn, bg=CORES['primary'])
        inner.pack(fill="x", padx=15, pady=10)

        # Ícone
        if icone and not is_admin_btn:
            caminho_icone = resource_path(f"Icones/{icone}")
            img = Image.open(caminho_icone)
            img = img.resize((24, 24), Image.LANCZOS)
            icon_img = ImageTk.PhotoImage(img)

            lbl_icon = tk.Label(inner, image=icon_img, bg=CORES['primary'])
            lbl_icon.image = icon_img
            lbl_icon.pack(side="left", padx=(0, 10))

        # Texto
        lbl_text = tk.Label(
            inner,
            text=texto,
            font=('Segoe UI', 10, 'bold') if is_admin_btn else ('Segoe UI', 10),
            bg=CORES['primary'],
            fg='white',
            anchor='w'
        )
        lbl_text.pack(side="left", fill="x")

        # Hover effects
        def on_enter(e):
            btn.config(bg='white')
            inner.config(bg='white')
            lbl_text.config(bg='white', fg=CORES['primary'])
            if icone and not is_admin_btn:
                lbl_icon.config(bg='white')

        def on_leave(e):
            btn.config(bg=CORES['primary'])
            inner.config(bg=CORES['primary'])
            lbl_text.config(bg=CORES['primary'], fg='white')
            if icone and not is_admin_btn:
                lbl_icon.config(bg=CORES['primary'])

        def on_click(e):
            comando()

        for widget in [btn, inner, lbl_text]:
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
            widget.bind("<Button-1>", on_click)

        if icone and not is_admin_btn:
            lbl_icon.bind("<Enter>", on_enter)
            lbl_icon.bind("<Leave>", on_leave)
            lbl_icon.bind("<Button-1>", on_click)

    
    def mostrar_home(self):
        home_frame = ttk.Frame(self.content_area, style='Main.TFrame')
        home_frame.pack(fill="both", expand=True, padx=50, pady=50)

        ttk.Label(
            home_frame,
            text="Bem-vindo ao Sistema Fiscal",
            font=('Segoe UI', 24, 'bold')
        ).pack(pady=(0, 10))

        ttk.Label(
            home_frame,
            text=f"Olá, {self.usuario_nome}! Selecione um módulo no menu lateral para começar.",
            font=('Segoe UI', 11)
        ).pack(pady=(0, 40))

        # =========================
        # AVISO DE ATUALIZAÇÃO
        # =========================
        if self.desatualizado:
            ttk.Label(
                home_frame,
                text=self.mensagem_update,
                foreground="#C0392B",
                font=('Segoe UI', 11, 'bold'),
                wraplength=600,     
                justify="center"
            ).pack(pady=(20, 10))

            btn_atualizar = ttk.Button(
                home_frame,
                text="Atualizar sistema",
                command=lambda: executar_atualizacao(self.dao)
            )

            if self.atualizacao_liberada:
                btn_atualizar.state(["!disabled"])
            else:
                btn_atualizar.state(["disabled"])

            btn_atualizar.pack(pady=5)



        # Cards informativos
        cards_container = ttk.Frame(home_frame, style='Main.TFrame')
        cards_container.pack(fill="both", expand=True)

        # Grid para os cards
        cards_container.grid_columnconfigure(0, weight=1)
        cards_container.grid_columnconfigure(1, weight=1)
        cards_container.grid_rowconfigure(0, weight=1)
        cards_container.grid_rowconfigure(1, weight=1)

        # Card 1
        self.criar_info_card(
            cards_container, 0, 0,
            "📊 Extração Automatizada",
            "Extraia dados fiscais de arquivos TXT e PDF diretamente para Excel com poucos cliques."
        )

        # Card 2
        self.criar_info_card(
            cards_container, 0, 1,
            "🔍 Comparação Inteligente",
            "Compare notas da SEFAZ com seu sistema e identifique divergências rapidamente."
        )

        # Card 3
        self.criar_info_card(
            cards_container, 1, 0,
            "📁 Triagem de Documentos",
            "Organize e mescle automaticamente PDFs de NF-e e CT-e a partir do SPED."
        )

        # Card 4
        self.criar_info_card(
            cards_container, 1, 1,
            "⚡ Gestão Eficiente",
            "Centralize toda gestão fiscal em um único sistema profissional e intuitivo."
        )

        # Rodapé
        footer = ttk.Frame(home_frame, style='Main.TFrame')
        footer.pack(fill="x", pady=(40, 0))
        
        ttk.Label(
            footer,
            text="v1.0 • MTSistem - Desenvolvido e licenciado por Miquéias Teles",
            font=('Segoe UI', 8),
            background=CORES['bg_main'],
            foreground=CORES['text_light']
        ).pack()

    def criar_info_card(self, parent, row, col, titulo, descricao):
        card = tk.Frame(
            parent,
            bg='white',
            highlightthickness=2,
            highlightbackground=CORES['bg_main']
        )
        card.grid(row=row, column=col, padx=15, pady=15, sticky="nsew")

        inner = tk.Frame(card, bg='white')
        inner.pack(fill="both", expand=True, padx=25, pady=25)

        tk.Label(
            inner,
            text=titulo,
            font=('Segoe UI', 13, 'bold'),
            bg='white',
            fg=CORES['primary'],
            anchor='w'
        ).pack(fill="x", pady=(0, 10))

        tk.Label(
            inner,
            text=descricao,
            font=('Segoe UI', 10),
            bg='white',
            fg=CORES['text_light'],
            anchor='w',
            wraplength=250,
            justify='left'
        ).pack(fill="x")

    def sair(self):
        resposta = messagebox.askyesno(
            "Sair do sistema",
            "Deseja sair e voltar para a tela de login?"
        )
        if not resposta:
            return

        self.root.destroy()

        root = tk.Tk()
        TelaLogin(root, versao_remota=self.versao_remota)
        root.mainloop()


# =========================
# VERSÃO EMBED DO EXTRATOR PDF
# =========================
class ExtratorFiscalPDFAppEmbed:
    def __init__(self, parent_frame, sistema_fiscal):
        self.parent_frame = parent_frame
        self.sistema_fiscal = sistema_fiscal
        self.arquivo_pdf = None
        self.criar_interface()

    def criar_interface(self):
        # Frame principal
        main_frame = ttk.Frame(self.parent_frame, style='Main.TFrame')
        main_frame.pack(fill="both", expand=True, padx=50, pady=30)

        # Header
        header_frame = ttk.Frame(main_frame, style='Main.TFrame')
        header_frame.pack(fill="x", pady=(0, 30))

        header_container = ttk.Frame(header_frame, style='Main.TFrame')
        header_container.pack(fill="x")

        # Ícone e título
        left_header = ttk.Frame(header_container, style='Main.TFrame')
        left_header.pack(side="left")

        caminho_icone = resource_path("Icones/pdf_azul.png")
        img = Image.open(caminho_icone)
        img = img.resize((40, 40), Image.LANCZOS)
        self.icon_header = ImageTk.PhotoImage(img)

        ttk.Label(
            left_header,
            image=self.icon_header,
            background=CORES['bg_main']
        ).pack(side="left", padx=(0, 15))

        title_frame = ttk.Frame(left_header, style='Main.TFrame')
        title_frame.pack(side="left")

        ttk.Label(
            title_frame,
            text="Extração PDF → Excel",
            font=('Segoe UI', 18, 'bold'),
            background=CORES['bg_main'],
            foreground=CORES['text_dark']
        ).pack(anchor="w")

        ttk.Label(
            title_frame,
            text="Extraia informações fiscais de relatórios em PDF",
            font=('Segoe UI', 9),
            background=CORES['bg_main'],
            foreground=CORES['text_light']
        ).pack(anchor="w")

        # Card de conteúdo
        content_container = ttk.Frame(main_frame, style='Main.TFrame')
        content_container.pack(fill="both", expand=True)

        center_frame = ttk.Frame(content_container, style='Main.TFrame')
        center_frame.pack(expand=True)

        card = ttk.Frame(center_frame, style='Card.TFrame', padding=40)
        card.pack(fill="both", expand=True, ipadx=100)

        ttk.Label(
            card,
            text="Configuração da Extração",
            font=('Segoe UI', 14, 'bold'),
            background=CORES['bg_card'],
            foreground=CORES['primary']
        ).pack(anchor="w", pady=(0, 25))

        self.lbl_pdf = self.criar_campo(
            card,
            "Arquivo PDF (.pdf)",
            "Selecione o relatório fiscal em PDF para processamento",
            self.selecionar_pdf
        )

        separator = ttk.Frame(card, height=2, style='Card.TFrame')
        separator.pack(fill="x", pady=30)

        btn_frame = ttk.Frame(card, style='Card.TFrame')
        btn_frame.pack(fill="x")

        ttk.Button(
            btn_frame,
            text="▶ Executar Extração e Exportar Excel",
            style='Primary.TButton',
            command=self.processar
        ).pack(fill="x", ipady=8)

    def criar_campo(self, parent, titulo, subtitulo, comando):
        campo_frame = ttk.Frame(parent, style='Card.TFrame')
        campo_frame.pack(fill="x")

        ttk.Label(
            campo_frame,
            text=titulo,
            font=('Segoe UI', 11, 'bold'),
            background=CORES['bg_card'],
            foreground=CORES['text_dark']
        ).pack(anchor="w", pady=(0, 5))

        ttk.Label(
            campo_frame,
            text=subtitulo,
            font=('Segoe UI', 9),
            background=CORES['bg_card'],
            foreground=CORES['text_light']
        ).pack(anchor="w", pady=(0, 12))

        input_frame = ttk.Frame(campo_frame, style='Card.TFrame')
        input_frame.pack(fill="x")

        entry = ttk.Entry(input_frame, font=('Segoe UI', 10))
        entry.pack(side="left", fill="x", expand=True, ipady=8)

        ttk.Button(
            input_frame,
            text="📁 Selecionar Arquivo",
            style='Secondary.TButton',
            command=comando
        ).pack(side="right", padx=(15, 0))

        return entry

    def selecionar_pdf(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo PDF",
            filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")]
        )
        if arquivo:
            self.lbl_pdf.delete(0, tk.END)
            self.lbl_pdf.insert(0, arquivo)

    def processar(self):
        caminho_pdf = self.lbl_pdf.get()
        if not caminho_pdf:
            messagebox.showwarning("Atenção", "Selecione um arquivo PDF.")
            return

        salvar_em = filedialog.asksaveasfilename(
            title="Salvar planilha como",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not salvar_em:
            return

        dados = []
        nota_atual = {
            "Fornecedor": None,
            "Data de Emissao": None,
            "Pedido": None,
            "Quantidade": 0.0
        }

        try:
            with pdfplumber.open(caminho_pdf) as pdf:
                for page in pdf.pages:
                    texto = page.extract_text()
                    if not texto:
                        continue

                    for linha in texto.split("\n"):
                        linha = linha.strip()

                        if linha.startswith("Fornecedor:"):
                            nota_atual["Fornecedor"] = linha.replace("Fornecedor:", "").strip()

                        if "Emissão:" in linha:
                            data = re.search(r"\d{2}/\d{2}/\d{4}", linha)
                            if data:
                                nota_atual["Data de Emissao"] = data.group()

                        # Linha de item começa com código numérico
                        if re.match(r"^\d{3,}", linha):
                            numeros = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha)

                            if numeros:
                                qtd = float(numeros[0].replace(".", "").replace(",", "."))
                                nota_atual["Quantidade"] += qtd



                        pedido_match = re.search(r"\b(\d{6})/\d{2}\b", linha)
                        if pedido_match and nota_atual["Pedido"] is None:
                            nota_atual["Pedido"] = pedido_match.group(1)

                        if linha.startswith("TOTAL DA NOTA:"):
                            numero_match = re.search(r"/\s*(\d+)", linha)
                            valor = float(linha.split()[-1].replace(".", "").replace(",", "."))

                            dados.append({
                                "Numero": numero_match.group(1) if numero_match else None,
                                "Pedido": nota_atual["Pedido"],
                                "Data de Emissao": nota_atual["Data de Emissao"],
                                "Fornecedor": nota_atual["Fornecedor"],
                                "Quantidade": nota_atual["Quantidade"],
                                "Valor": valor
                            })

                            nota_atual = {
                                "Fornecedor": None,
                                "Data de Emissao": None,
                                "Pedido": None,
                                "Quantidade": 0.0
                            }

            if not dados:
                messagebox.showwarning("Aviso", "Nenhuma nota encontrada no arquivo PDF.")
                return

            df = pd.DataFrame(dados)
            df["Data de Emissao"] = pd.to_datetime(df["Data de Emissao"], dayfirst=True, errors="coerce")
            df.to_excel(salvar_em, index=False)

            messagebox.showinfo(
                "Sucesso", 
                f"Extração concluída com sucesso!\n\nNotas extraídas: {len(df)}\nArquivo salvo em:\n{salvar_em}"
            )

        except Exception as e:
            messagebox.showerror("Erro ao processar", f"Ocorreu um erro durante a extração:\n\n{str(e)}")

# =====================================================
# TELA DE CONFIGURAÇÕES DO SISTEMA (ADMIN)
# =====================================================

class TelaConfiguracoesSistema:
    def __init__(self, parent):
        self.dao = UsuarioDAO()

        self.janela = tk.Toplevel(parent)
        self.janela.title("Configurações do Sistema")
        self.janela.geometry("720x600")
        self.janela.configure(bg=CORES['bg_main'])

        caminho_icone = resource_path("Icones/logo.ico")
        self.janela.iconbitmap(caminho_icone)

        self.centralizar()
        self.criar_interface()

    def centralizar(self):
        self.janela.update_idletasks()
        w = self.janela.winfo_width()
        h = self.janela.winfo_height()
        x = (self.janela.winfo_screenwidth() - w) // 2
        y = (self.janela.winfo_screenheight() - h) // 2
        self.janela.geometry(f"{w}x{h}+{x}+{y}")

    def criar_interface(self):
        container = ttk.Frame(self.janela, padding=30, style="Main.TFrame")
        container.pack(fill="both", expand=True)

        ttk.Label(
            container,
            text="Configurações do Sistema",
            font=('Segoe UI', 18, 'bold')
        ).pack(anchor="w", pady=(0, 25))

        # ==================================================
        # CARD – CONTROLE DE ATUALIZAÇÃO
        # ==================================================
        card = ttk.Frame(container, padding=20, style="Card.TFrame")
        card.pack(fill="x")

        ttk.Label(
            card,
            text="Atualização e Controle do Sistema",
            font=('Segoe UI', 14, 'bold')
        ).pack(anchor="w", pady=(0, 15))

        # -------- Versão atual no servidor --------
        versao_atual = self.dao.get_config("versao_atual", VERSAO_ATUAL)

        ttk.Label(card, text="Versão liberada no servidor").pack(anchor="w")
        self.entry_versao = ttk.Entry(card)
        self.entry_versao.insert(0, versao_atual)
        self.entry_versao.pack(fill="x", pady=(0, 10))

        # -------- Caminho do EXE --------
        exe = self.dao.get_config("exe_atualizacao", "")
        ttk.Label(card, text="Executável de atualização").pack(anchor="w")
        self.entry_exe = ttk.Entry(card)
        self.entry_exe.insert(0, exe)
        self.entry_exe.pack(fill="x", pady=(0, 10))

        # -------- Mensagem --------
        mensagem = self.dao.get_config("mensagem_update", "")
        ttk.Label(card, text="Mensagem para os usuários").pack(anchor="w")
        self.txt_mensagem = tk.Text(card, height=4)
        self.txt_mensagem.insert("1.0", mensagem)
        self.txt_mensagem.pack(fill="x", pady=(0, 10))

        # -------- Flags --------
        liberar = self.dao.get_config("atualizacao_liberada", "NAO")
        bloqueado = self.dao.get_config("sistema_bloqueado", "NAO")

        self.var_liberar = tk.IntVar(value=1 if liberar == "SIM" else 0)
        self.var_bloquear = tk.IntVar(value=1 if bloqueado == "SIM" else 0)

        ttk.Checkbutton(
            card,
            text="Liberar atualização para usuários",
            variable=self.var_liberar
        ).pack(anchor="w")

        ttk.Checkbutton(
            card,
            text="Bloquear acesso ao sistema (exceto admin)",
            variable=self.var_bloquear
        ).pack(anchor="w", pady=(5, 15))

        ttk.Button(
            card,
            text="Salvar configurações",
            style="Primary.TButton",
            command=self.salvar
        ).pack(anchor="e")

    def salvar(self):
        self.dao.set_config("versao_atual", self.entry_versao.get().strip())
        self.dao.set_config("exe_atualizacao", self.entry_exe.get().strip())
        self.dao.set_config(
            "mensagem_update",
            self.txt_mensagem.get("1.0", "end").strip()
        )
        self.dao.set_config(
            "atualizacao_liberada",
            "SIM" if self.var_liberar.get() else "NAO"
        )
        self.dao.set_config(
            "sistema_bloqueado",
            "SIM" if self.var_bloquear.get() else "NAO"
        )

        messagebox.showinfo(
            "Sucesso",
            "Configurações salvas com sucesso."
        )



if __name__ == "__main__":
    versao_remota = verificar_versao_no_startup()
    root = tk.Tk()
    TelaLogin(root, versao_remota)
    root.mainloop()